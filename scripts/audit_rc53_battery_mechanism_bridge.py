#!/usr/bin/env python3
"""Audit RC53 source metadata and archive directories without reading summary values."""

from __future__ import annotations

import hashlib
import io
import json
import pathlib
import urllib.request
import zipfile

from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).resolve().parents[1]
OUT = ROOT / "research" / "reproducibility" / "rc53-battery-mechanism-bridge-schema-audit.json"
API = "https://zenodo.org/api/records/10637534"


def fetch(url: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": "UnsolvedProblems-RC53/1.0"})
    with urllib.request.urlopen(request, timeout=120) as response:
        return response.read()


def main() -> None:
    record = json.loads(fetch(API))
    files = {item["key"]: item for item in record["files"]}
    metadata_bytes = fetch(files["experiment_metadata.xlsx"]["links"]["self"])
    metadata_md5 = hashlib.md5(metadata_bytes).hexdigest()
    workbook = load_workbook(io.BytesIO(metadata_bytes), data_only=True, read_only=True)
    rows = []
    for sheet in workbook.worksheets:
        iterator = sheet.iter_rows(values_only=True)
        header = [str(value) for value in next(iterator)]
        for values in iterator:
            row = dict(zip(header, values))
            rows.append({
                "experiment": str(row["Expt"]),
                "cell": str(row["Cell"]),
                "temperatureC": int(row["Temp"]),
                "socRange": str(row["SoC range"]),
                "ageSets": int(row["Age sets"]),
            })

    # Archive directory counts were obtained through HTTP range requests. The audit
    # deliberately records only member names, never opens a member stream.
    import fsspec

    archive_audit = []
    for key, item in files.items():
        if not key.endswith(".zip"):
            continue
        with fsspec.open(item["links"]["self"], "rb", block_size=1_048_576).open() as remote:
            with zipfile.ZipFile(remote) as archive:
                names = archive.namelist()
        performance = [name for name in names if "/Summary Data/Performance Summary/" in name and name.endswith("Processed Data.csv")]
        archive_audit.append({
            "file": key,
            "bytes": item["size"],
            "md5": item["checksum"].removeprefix("md5:"),
            "memberCount": len(names),
            "performanceSummaryCount": len(performance),
            "performanceSummaryMembers": sorted(performance),
            "summaryValuesRead": False,
        })

    output = {
        "auditId": "RC53-BATTERY-MECHANISM-BRIDGE-SCHEMA-AUDIT-0.1",
        "cycleId": "RC-2026-53",
        "record": "https://zenodo.org/records/10637534",
        "metadata": {
            "rows": len(rows),
            "sheetNames": workbook.sheetnames,
            "md5": metadata_md5,
            "registeredMd5": files["experiment_metadata.xlsx"]["checksum"].removeprefix("md5:"),
            "matches": metadata_md5 == files["experiment_metadata.xlsx"]["checksum"].removeprefix("md5:"),
            "cells": rows,
        },
        "archives": archive_audit,
        "totals": {
            "metadataRows": len(rows),
            "performanceSummaryMembers": sum(item["performanceSummaryCount"] for item in archive_audit),
            "pilotCells": 4,
            "developmentCells": 14,
            "untouchedTargetCells": 22,
        },
        "summaryValuesRead": False,
    }
    OUT.write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    print(f"RC53 schema audit: {len(rows)} metadata rows, {output['totals']['performanceSummaryMembers']} performance summaries, no summary values read.")


if __name__ == "__main__":
    main()
